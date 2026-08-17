"""Standalone Shimanami Kaido workbook — the three ways to do the crossing."""
import os, sys, datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .itineraries import SHIMANAMI_VARIANTS, SHIMANAMI_LABELS, SHIMANAMI_VERDICT

TAB_NAME = {"nobike": "No cycling", "1day": "1 day by bus",
            "2day": "2 days riding", "3day": "3 days riding"}
from .refs import JPY_ILS, JPY_ILS_ASOF, LIVE_FORMULA, RATE_HELP, SOURCES, resolve, places_in, gmap, price_unit, PRICE_NOTE, PRICE_BODY
from . import build_sheets as MS

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "out", "sheets")

EVIDENCE = [
 ("How long is it, really?", "~80 km, not 70",
  "A GPS-measured log puts the full ride at 79.8 km — the bridge approach spirals and the "
  "run into Imabari station add about 10 km to the number everyone quotes.", "train_cycling"),
 ("How fast do normal people ride it?", "10 km/h, ~8 hours",
  "The rental operator's own figure for recreational riders on rental bikes, including an "
  "hour for lunch. Experienced riders on road bikes: 3–4 hours.", "shimanami_op"),
 ("Does the operator think one day is enough?", "No",
  "Shimanami Japan explicitly says the route need not be done in one day and encourages "
  "two or three.", "shimanami_op"),
 ("What do casual riders recommend?", "3 days",
  "Epic Road Rides recommends three days for casual riders. A local specialist says "
  "'at least two days'.", "epic_roads"),
 ("Did anyone say two days was too slow?", "Nobody",
  "Not one source. Two independent accounts said two days felt RUSHED — one took three "
  "days and still wished they'd gone slower, missing Kōsanji and the salt factory.", "touring_shim"),
 ("Which direction?", "Imabari → Onomichi",
  "Our direction is the right one. On strong-wind days the prevailing westerlies mean "
  "Imabari→Onomichi gets the tailwind, and both of the route's real hills (Ōshima's "
  "Miyakubo ~70 m and Taura ~78 m) come in the first 20 km while you're fresh.", "touring_wind"),
 ("Where is the true midpoint?", "Ōmishima, not Setoda",
  "Setoda is ~38 km from Onomichi, so from Imabari it is the FAR side of halfway. "
  "Sleeping on Ōmishima splits it 40/40; sleeping at Setoda makes day 1 much harder.", "touring_dist"),
 ("Can we use e-bikes?", "NO — this is the trap",
  "Electric-assist bikes and e-bikes MUST be returned to the terminal they were rented "
  "from, so they cannot do a one-way crossing. They are also one-day rentals only. "
  "Cross bike ¥3,000/day + ¥1,000 one-way drop-off is the only option that works.", "shimanami_cyc"),
 ("What about the bags?", "¥2,200 a bag, same day",
  "Sagawa's Hands-Free Cycling moves luggage between Onomichi, the islands and Imabari the "
  "SAME day. Drop by 09:00 in Onomichi, 10:00 on the islands and at Imabari. Reserve the "
  "night before.", "sagawa_bags"),
 ("Daylight in mid-October?", "~11 hours, unlit bridges",
  "Sunset falls around 17:20–17:40 and the bridge cycleways have no lighting. A beginner "
  "needing 8–10 hours has almost no margin for a puncture or a wrong turn.", "touring_shim"),
 ("Is October a good time?", "Yes — one of the two best",
  "April–May and October–November are named as the optimal windows. Typhoon risk has "
  "largely subsided by October and autumn is a relatively weak-wind season.", "touring_wind"),
 ("The built-in escape valve", "Setoda → Onomichi by boat",
  "Setouchi Cruising takes you Setoda→Onomichi in 40 min for ¥1,500 plus ¥500 for the "
  "bike, which rolls straight aboard. Sailings 11:25 / 13:25 / 15:00 / 17:00. Plan it as "
  "an option, not a failure.", "setouchi_cru"),
]

def build():
    wb = Workbook(); wb.remove(wb.active)

    # one tab per variant
    for v, days in SHIMANAMI_VARIANTS.items():
        ws = wb.create_sheet(TAB_NAME[v])
        ws.sheet_properties.tabColor = {"nobike": "6B7481", "1day": "8A9099", "2day": "3E6B8A", "3day": "26624A"}[v]
        MS._title_block(ws, f"Shimanami Kaido — {SHIMANAMI_LABELS[v]}", SHIMANAMI_VERDICT[v], 7)
        MS._header(ws, 4, ["Day", "What you'll do", "Getting there", "Watch out for",
                           "Sleep in", "Legs", "Map links"], wraps=(2, 3, 4, 5, 6, 7))
        MS._w(ws, [30, 72, 40, 60, 34, 52, 30])
        r = 5
        for i, d in enumerate(days):
            ws.cell(r, 1, d["title"]).font = Font(name=MS.SANS, size=11, bold=True, color=MS.INK)
            ws.cell(r, 2, MS._bullets(d["do"]))
            ws.cell(r, 3, d["travel"]).font = Font(name=MS.SANS, size=10, color=MS.MUTED)
            ws.cell(r, 4, MS._bullets(d["watch"])).font = Font(name=MS.SANS, size=9.5, color=MS.ACCENT)
            ws.cell(r, 5, d["sleep"]).font = Font(name=MS.SANS, size=10, bold=True)
            ws.cell(r, 6, "\n".join(
                f"{a} → {b}  {dep}  {how}" + (f"  ¥{yen:,}" if yen else "")
                for a, b, dep, arr, how, yen in d["legs"]))
            ws.cell(r, 6).font = Font(name=MS.MONO, size=8.5, color=MS.INK)
            blob = " ".join([d["title"]] + d["do"] + [d["travel"]] + d["watch"] + [d["sleep"]])
            ws.cell(r, 7, "\n".join(n for n, _ in places_in(blob)))
            ws.cell(r, 7).font = Font(name=MS.SANS, size=9, color=MS.MUTED)
            MS._band(ws, r, 7, i % 2 == 1)
            for c in range(1, 8):
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(r, c).border = MS.GRID
            ws.row_dimensions[r].height = min(300, 15 * max(len(d["do"]) * 3, len(d["watch"]) * 3, 8))
            r += 1
        yen = sum(l[5] for d in days for l in d["legs"] if l[5] is not None)
        ws.cell(r + 1, 5, "Transport + rental, sourced").font = Font(name=MS.SANS, bold=True)
        c = ws.cell(r + 1, 6, yen); c.number_format = '"¥"#,##0'
        c.font = Font(name=MS.MONO, size=11, bold=True)
        ws.sheet_view.showGridLines = False

    # ── Transport tab: every leg of every variant, one table ──────────────
    wt = wb.create_sheet("Transport"); wt.sheet_properties.tabColor = "5B6672"
    n = 10
    MS._title_block(wt, "Shimanami Kaido — every leg, all four ways",
                    "ALL PRICES ARE PER PERSON. " + PRICE_BODY, n)
    wt.cell(4, 1, "Live rate  JPY → NIS").font = Font(name=MS.SANS, size=9.5, bold=True)
    rate = wt.cell(4, 2, LIVE_FORMULA)
    rate.font = Font(name=MS.MONO, size=11, bold=True, color=MS.ACCENT)
    rate.number_format = "0.00000"
    rate.fill = PatternFill("solid", fgColor=MS.ACC_BG)
    rate.alignment = Alignment(horizontal="center")
    wt.cell(3, 3, "Rate below is live in Google Sheets, fixed elsewhere — see the note")\
      .font = Font(name=MS.SANS, size=8.5, italic=True, color=MS.MUTED)
    wt.row_dimensions[3].height = 13
    wt.cell(4, 3, RATE_HELP).font = Font(name=MS.SANS, size=9, italic=True, color=MS.MUTED)
    wt.row_dimensions[5].height = 6
    MS._header(wt, 6, ["Variant", "Day", "From", "To", "Departure", "Arrival",
                       "How you travel", "Price (Yen)\nPER PERSON", "Charged",
                       "Where the price comes from"], wraps=(7, 8, 10))
    MS._w(wt, [16, 30, 26, 26, 24, 20, 40, 14, 13, 46])
    tier_colour = {"operator": "FF26624A", "official": "FF26624A",
                   "third-party": "FF8A6410", "traveller": "FF8A6410"}
    r, band = 7, 0
    for v, days in SHIMANAMI_VARIANTS.items():
        for d in days:
            for a, b, dep, arr, how, yen in d["legs"]:
                wt.cell(r, 1, TAB_NAME[v]).font = Font(name=MS.SANS, size=9.5, bold=True, color=MS.ACCENT)
                wt.cell(r, 2, d["title"]).font = Font(name=MS.SANS, size=9, color=MS.MUTED)
                wt.cell(r, 3, a).font = Font(name=MS.SANS, size=9.5, bold=True)
                wt.cell(r, 4, b).font = Font(name=MS.SANS, size=9.5, bold=True)
                wt.cell(r, 5, dep).font = Font(name=MS.MONO, size=9, color=MS.MUTED)
                wt.cell(r, 6, arr).font = Font(name=MS.MONO, size=9, color=MS.MUTED)
                wt.cell(r, 7, how).font = Font(name=MS.SANS, size=9.5)
                if yen is not None:
                    c = wt.cell(r, 8, yen); c.number_format = '"¥"#,##0'
                    c.font = Font(name=MS.MONO, size=9.5)
                    c.alignment = Alignment(horizontal="right", indent=1)
                    u = price_unit(how)
                    wt.cell(r, 9, u).font = Font(name=MS.SANS, size=9,
                        bold=(u == "per bag"), color=(MS.ACCENT if u == "per bag" else MS.MUTED))
                    src = resolve(how)
                    if src:
                        MS._link(wt, r, 10, src[0], src[1], size=9)
                        wt.cell(r, 10).alignment = Alignment(wrap_text=True, vertical="center", indent=1)
                    else:
                        wt.cell(r, 10, "unsourced — verify").font = Font(
                            name=MS.SANS, size=9, italic=True, color=MS.ACCENT)
                else:
                    wt.cell(r, 10, "no published figure").font = Font(
                        name=MS.SANS, size=9, italic=True, color=MS.MUTED)
                MS._band(wt, r, n, band % 2 == 1)
                for c in range(1, n + 1):
                    wt.cell(r, c).border = MS.GRID
                    if c != 10:
                        wt.cell(r, c).alignment = Alignment(
                            vertical="center", wrap_text=(c == 7),
                            horizontal=("right" if c == 8 else None),
                            indent=(1 if c == 8 else 0))
                wt.row_dimensions[r].height = 26
                r += 1; band += 1
    last = r - 1
    for i, (v, days) in enumerate(SHIMANAMI_VARIANTS.items()):
        tot = sum(l[5] for d in days for l in d["legs"] if l[5] is not None)
        wt.cell(last + 2 + i, 7, f"TOTAL — {SHIMANAMI_LABELS[v]}, one person")
        wt.cell(last + 2 + i, 7).font = Font(name=MS.SANS, size=10, bold=True)
        wt.cell(last + 2 + i, 7).alignment = Alignment(horizontal="right")
        c = wt.cell(last + 2 + i, 8, tot); c.number_format = '"¥"#,##0'
        c.font = Font(name=MS.MONO, size=10.5, bold=True, color=MS.INK)
        c.fill = PatternFill("solid", fgColor=MS.ACC_BG)
    wt.auto_filter.ref = f"A6:{get_column_letter(n)}{last}"
    wt.sheet_view.showGridLines = False

    # the evidence tab
    ws = wb.create_sheet("Why 2–3 days"); ws.sheet_properties.tabColor = "B08A3E"
    MS._title_block(ws, "How many days does the Shimanami Kaido actually need?",
                    "Every figure below is from the operator, a measured GPS log, or a named "
                    "rider. Nothing is estimated.", 4)
    MS._header(ws, 4, ["Question", "Short answer", "The evidence", "Source"], wraps=(3,))
    MS._w(ws, [40, 30, 86, 50])
    r = 5
    for i, (q, a, ev, key) in enumerate(EVIDENCE):
        ws.cell(r, 1, q).font = Font(name=MS.SANS, size=10.5, bold=True, color=MS.INK)
        ws.cell(r, 2, a).font = Font(name=MS.SANS, size=10.5, bold=True, color=MS.ACCENT)
        ws.cell(r, 3, ev).font = Font(name=MS.SANS, size=10, color=MS.INK)
        label, url, tier = SOURCES[key]
        MS._link(ws, r, 4, label, url, size=9)
        MS._band(ws, r, 4, i % 2 == 1)
        for c in range(1, 5):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = MS.GRID
        ws.row_dimensions[r].height = max(40, 12 * (len(ev) // 80 + 2))
        r += 1
    ws.sheet_view.showGridLines = False

    ws = wb.create_sheet("Sources"); ws.sheet_properties.tabColor = "5A5A5A"
    MS.build_sources(ws)

    wb.active = 0
    os.makedirs(OUT, exist_ok=True)
    path = os.path.join(OUT, "Shimanami-Kaido.xlsx")
    wb.save(path)
    print(f"Shimanami-Kaido.xlsx: {len(SHIMANAMI_VARIANTS)} variants + evidence tab "
          f"({len(EVIDENCE)} findings)")

if __name__ == "__main__":
    build()
