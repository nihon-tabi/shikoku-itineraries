"""Generate the .xlsx workbooks and the CSVs from src/itineraries.py.

Run from the project root:  python -m src.build_sheets

Run:  venv/bin/python build/make_sheets.py
"""
import csv, datetime, os, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .itineraries import OPTIONS, GLOSSARY, FOLIAGE, RATE_NOTE
from .refs import JPY_ILS, JPY_ILS_ASOF, LIVE_FORMULA, RATE_HELP, ATTRACTIONS, BOOKINGS, SOURCES, resolve, gmap, ADMISSIONS, DWELL, TIDES, day_places, pin, TRAILHEADS, walks_for, price_unit, PRICE_NOTE, PRICE_BODY

LINK = "FF1F5FA8"

def _link(ws, r, c, text, url, size=9.5, bold=False):
    """Write a clickable link.

    Uses the =HYPERLINK() FORMULA rather than openpyxl's cell.hyperlink. The
    relationship-based form openpyxl writes declares xmlns:r inline on each
    <hyperlink> element instead of on the worksheet root, which Excel tolerates
    but Google Sheets and Apple Numbers frequently drop on import. The formula
    survives every reader, including xlsx -> Google Sheets conversion.
    """
    esc_t = str(text).replace('"', '""')
    esc_u = str(url).replace('"', '""')
    cell = ws.cell(r, c, f'=HYPERLINK("{esc_u}","{esc_t}")')
    cell.font = Font(name=SANS, size=size, bold=bold, color=LINK, underline="single")
    return cell

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "out", "sheets")

# ── palette (cool washi + sumi ink + one vermilion, matching the dashboard) ──
INK      = "FF14171D"
MUTED    = "FF6B7481"
PAPER    = "FFFFFFFF"
BAND     = "FFF7F8FA"
HEAD_BG  = "FF1B2129"
HEAD_FG  = "FFF2F4F7"
ACCENT   = "FFC0392B"
ACC_BG   = "FFFBEDEB"
OK_BG    = "FFEDF5F0"
TAB_A    = "1B2129"
TAB_B    = "5B6672"

HAIR  = Side(style="thin", color="FFE1E5EA")
GRID  = Border(bottom=HAIR)
SANS  = "Calibri"
MONO  = "Consolas"

def _w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _title_block(ws, text, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, text)
    c.font = Font(name=SANS, size=15, bold=True, color=INK)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, sub)
    c.font = Font(name=SANS, size=9.5, italic=True, color=MUTED)
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

def _header(ws, row, headers, wraps=()):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = Font(name=SANS, size=9.5, bold=True, color=HEAD_FG)
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(vertical="center", horizontal="left",
                                wrap_text=(i in wraps))
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row + 1, 1)

def _bullets(items):
    return "\n".join(f"•  {t}" for t in items)


def _highlights(d):
    """A day whose ORDER is forced renders as a numbered sequence, not bullets.

    Days without a declared flow are genuinely pick-your-own and stay as a list —
    the distinction is the point, so the traveller can tell at a glance which
    days they can shuffle and which they cannot.
    """
    flow = d.get("flow")
    if not flow:
        return _bullets(d["do"])
    return ("THE ORDER MATTERS TODAY \u2014 do it in this sequence:\n\n"
            + "\n\n".join(f"{i}.  [{w}]  {x}" for i, (w, x) in enumerate(flow, 1)))

def _band(ws, r, ncols, on):
    if on:
        for c in range(1, ncols + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=BAND)

# ── Itinerary tab ───────────────────────────────────────────────────────────
# NOTE: no "map links" column. A spreadsheet cell can hold exactly ONE hyperlink,
# so a day with nine places cannot carry nine pins in one cell — the old column
# was a list of names with nothing clickable in it. The pins live on the
# "Places & timings" tab instead, one per row, every one live, filterable by date.
ITIN_HEAD = ["Date", "Day", "Where / what the day is",
             "☎ BOOK BEFORE THE DAY",
             "Highlights — what you'll actually do",
             "Getting there", "Watch out for", "Sleep in"]

def itinerary_rows(opt):
    rows = []
    for d in opt["days"]:
        dt = datetime.date.fromisoformat(d["date"])
        rows.append([d["date"], dt.strftime("%a"), d["title"],
                     _bullets(d.get("book", [])),
                     _highlights(d), d["travel"],
                     _bullets(d["watch"]) + (
                         "\n\n\u25b6 THIS DAY HAS %d ALTERNATIVES \u2014 see the "
                         "\u201cDay alternatives\u201d tab." % (len(d["alts"]) + 1)
                         if d.get("alts") else ""), d["sleep"]])
    return rows

def build_itinerary(ws, opt):
    n = len(ITIN_HEAD)
    _title_block(ws, opt["name"], opt["verdict"] +
                 "   \u2014   Every place named here is a live pin on the "
                 "\u201cPlaces & timings\u201d tab: one per row, filterable by date. Walks and their trailheads are on \u201cWalks & trailheads\u201d. "
                 "A spreadsheet cell can only hold one link, so they cannot all sit here.", n)
    _header(ws, 4, ITIN_HEAD, wraps=(3, 4, 5, 6, 7, 8))
    _w(ws, [11, 6, 30, 46, 74, 42, 52, 26])
    r = 5
    for i, row in enumerate(itinerary_rows(opt)):
        for c, v in enumerate(row[:8], 1):
            ws.cell(r, c, v)
        if row[3]:                           # the booking column, when it has content
            ws.cell(r, 4).font = Font(name=SANS, size=10, bold=True, color=ACCENT)
            ws.cell(r, 4).fill = PatternFill("solid", fgColor="FFF6E9E7")
        u = pin(row[7])                      # "Sleep in" -> the actual bed
        if u:
            _link(ws, r, 8, row[7], u, size=9.5)

        _band(ws, r, n, i % 2 == 1)
        ws.cell(r, 1).font = Font(name=MONO, size=9.5, color=INK)
        ws.cell(r, 2).font = Font(name=MONO, size=9.5, color=MUTED)
        ws.cell(r, 3).font = Font(name=SANS, size=10.5, bold=True, color=INK)
        ws.cell(r, 4).font = Font(name=SANS, size=10, color=INK)
        ws.cell(r, 5).font = Font(name=SANS, size=10, color=MUTED)
        ws.cell(r, 6).font = Font(name=SANS, size=9.5, color=ACCENT)
        ws.cell(r, 7).font = Font(name=SANS, size=10, bold=True, color=INK)
        for c in range(1, n + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = GRID
        lines = max(len(row[3].split("\n")) * 2.6, len(row[5].split("\n")) * 2.2, 6)
        ws.row_dimensions[r].height = min(240, 15 * lines)
        r += 1
    ws.auto_filter.ref = f"A4:{get_column_letter(n)}{r-1}"
    ws.sheet_view.showGridLines = False

# ── Transport tab ───────────────────────────────────────────────────────────
TR_HEAD = ["Date", "From", "To", "Departure time", "Arrival time",
           "How you travel", "Price (Yen)\nPER PERSON", "Price (NIS)\nPER PERSON",
           "Charged", "Where the price comes from", "How solid"]

def transport_rows(opt):
    rows = []
    for d in opt["days"]:
        for a, b, dep, arr, how, yen in d["legs"]:
            rows.append([d["date"], a, b, dep, arr, how, yen])
    return rows

def build_transport(ws, opt):
    n = len(TR_HEAD)
    _title_block(ws, f"{opt['name']}  —  every leg, with real published times",
                 "ALL PRICES ARE PER PERSON. " + PRICE_BODY +
                 " Blank price = the operator publishes no figure; nothing here is estimated.", n)

    ws.cell(4, 1, "Live rate  JPY → NIS").font = Font(name=SANS, size=9.5, bold=True)
    rate = ws.cell(4, 2, LIVE_FORMULA)
    rate.font = Font(name=MONO, size=11, bold=True, color=ACCENT)
    rate.number_format = "0.00000"
    rate.fill = PatternFill("solid", fgColor=ACC_BG)
    rate.alignment = Alignment(horizontal="center")
    ws.cell(3, 3, "Rate below is live in Google Sheets, fixed elsewhere — see the note")\
      .font = Font(name=SANS, size=8.5, italic=True, color=MUTED)
    ws.row_dimensions[3].height = 13
    ws.cell(4, 3, RATE_HELP).font = Font(name=SANS, size=9, italic=True, color=MUTED)
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 6

    _header(ws, 6, TR_HEAD, wraps=(6, 7, 8, 10))
    _w(ws, [11, 26, 26, 20, 20, 40, 14, 15, 13, 44, 12])

    tier_colour = {"operator": "FF26624A", "official": "FF26624A",
                   "third-party": "FF8A6410", "traveller": "FF8A6410"}
    r = 7
    for i, row in enumerate(transport_rows(opt)):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v if v is not None else "")
        # From / To become pins where the label names a real place. "up and back"
        # and "Big bags" resolve to nothing and stay plain, deliberately.
        for c in (2, 3):
            u = pin(row[c - 1])
            if u:
                _link(ws, r, c, row[c - 1], u, size=9.5)
        ws.cell(r, 8, f"=IF(G{r}=\"\",\"\",ROUND(G{r}*$B$4,2))")
        _band(ws, r, n, i % 2 == 1)
        for c in range(1, 9):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 6))
            cell.font = Font(name=(MONO if c in (1, 4, 5, 7, 8) else SANS),
                             size=9.5,
                             bold=(c in (2, 3)),
                             color=(MUTED if c in (4, 5) else INK))
        ws.cell(r, 7).number_format = '"¥"#,##0'
        ws.cell(r, 8).number_format = '"₪"#,##0.00'
        # breathing room so the money doesn't run into the source column
        ws.cell(r, 7).alignment = Alignment(vertical="center", horizontal="right", indent=1)
        ws.cell(r, 8).alignment = Alignment(vertical="center", horizontal="right", indent=2)

        unit = price_unit(row[5]) if row[6] is not None else ""
        u = ws.cell(r, 9, unit)
        u.font = Font(name=SANS, size=9, bold=(unit == "per bag"),
                      color=(ACCENT if unit == "per bag" else MUTED))
        u.alignment = Alignment(vertical="center", horizontal="center")
        src = resolve(row[5])
        if row[6] is None:
            ws.cell(r, 10, "no published figure found — left blank on purpose")
            ws.cell(r, 10).font = Font(name=SANS, size=9, italic=True, color=MUTED)
            ws.cell(r, 11, "—").font = Font(name=SANS, size=9, color=MUTED)
        elif src:
            label, url, tier = src
            _link(ws, r, 10, label, url, size=9)
            ws.cell(r, 10).alignment = Alignment(vertical="center", wrap_text=True, indent=1)
            t = ws.cell(r, 11, tier)
            t.font = Font(name=SANS, size=9, bold=True, color=tier_colour[tier])
        else:
            ws.cell(r, 10, "unsourced — verify before you rely on it")
            ws.cell(r, 10).font = Font(name=SANS, size=9, italic=True, color=ACCENT)
            ws.cell(r, 11, "check").font = Font(name=SANS, size=9, bold=True, color=ACCENT)
        for c in range(1, n + 1):
            ws.cell(r, c).border = GRID
        ws.row_dimensions[r].height = 28
        r += 1

    last = r - 1
    tr = r + 1
    ws.cell(tr, 6, "TOTAL, ONE PERSON — sourced legs only").font = Font(name=SANS, size=10.5, bold=True)
    ws.cell(tr, 6).alignment = Alignment(horizontal="right")
    for col, fmt in ((7, '"¥"#,##0'), (8, '"₪"#,##0.00')):
        c = ws.cell(tr, col, f"=SUM({get_column_letter(col)}7:{get_column_letter(col)}{last})")
        c.font = Font(name=MONO, size=11, bold=True, color=INK)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=ACC_BG)
        c.border = Border(top=Side(style="medium", color=INK))
    ws.cell(tr + 1, 6, "TOTAL, TWO PEOPLE").font = Font(name=SANS, size=10.5, bold=True, color=ACCENT)
    ws.cell(tr + 1, 6).alignment = Alignment(horizontal="right")
    for col, fmt in ((7, '"¥"#,##0'), (8, '"₪"#,##0.00')):
        c = ws.cell(tr + 1, col, f"={get_column_letter(col)}{tr}*2")
        c.font = Font(name=MONO, size=11, bold=True, color=ACCENT)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor=ACC_BG)
    ws.cell(tr + 3, 6, PRICE_NOTE).font = Font(name=SANS, size=9, italic=True, color=MUTED)
    ws.merge_cells(start_row=tr + 3, start_column=6, end_row=tr + 3, end_column=11)
    ws.cell(tr + 3, 6).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[tr + 3].height = 46
    ws.cell(tr + 5, 6, "Excludes beds, food and admissions. Treat as a floor, not a budget.")\
      .font = Font(name=SANS, size=9, italic=True, color=MUTED)
    ws.auto_filter.ref = f"A6:{get_column_letter(n)}{last}"
    ws.sheet_view.showGridLines = False

# ── Glossary tab ────────────────────────────────────────────────────────────
def build_glossary(ws):
    head = ["Term", "Japanese", "What it means", "Why it matters on this trip"]
    _title_block(ws, "Glossary — the words used in this plan",
                 "Everything assumed elsewhere, spelled out.", 4)
    _header(ws, 4, head, wraps=(3, 4))
    _w(ws, [26, 18, 66, 72])
    r = 5
    for i, (term, jp, what, why) in enumerate(GLOSSARY):
        for c, v in enumerate((term, jp, what, why), 1):
            ws.cell(r, c, v)
        _band(ws, r, 4, i % 2 == 1)
        ws.cell(r, 1).font = Font(name=SANS, size=10.5, bold=True, color=INK)
        ws.cell(r, 2).font = Font(name=SANS, size=10.5, color=MUTED)
        ws.cell(r, 3).font = Font(name=SANS, size=10, color=INK)
        ws.cell(r, 4).font = Font(name=SANS, size=10, color=ACCENT)
        for c in range(1, 5):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = GRID
        ws.row_dimensions[r].height = max(46, 13 * (len(why) // 60 + len(what) // 55 + 1))
        r += 1
    ws.sheet_view.showGridLines = False

# ── Foliage tab ─────────────────────────────────────────────────────────────
def build_foliage(ws):
    head = ["Spot", "Elevation", "Best viewing starts", "Peak", "Ends",
            "For 9–24 Oct 2026", "What that means"]
    _title_block(ws, "Autumn foliage — when each place actually turns",
                 "Dates are the 2025 season from Nihon Kishou's own data. "
                 "Year-to-year spread is only about 7–8 days.", 7)
    _header(ws, 4, head, wraps=(7,))
    _w(ws, [30, 13, 19, 13, 12, 17, 62])
    r = 5
    for i, row in enumerate(FOLIAGE):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, v)
        _band(ws, r, 7, i % 2 == 1)
        good = row[5].startswith("ON TIME")
        marg = row[5].startswith("MARGINAL")
        ws.cell(r, 1).font = Font(name=SANS, size=10.5, bold=True, color=INK)
        ws.cell(r, 6).font = Font(name=SANS, size=10, bold=True,
                                  color=("FF26624A" if good else ACCENT))
        if good or marg:
            for c in range(1, 8):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=OK_BG)
        for c in range(2, 8):
            if c != 6:
                ws.cell(r, c).font = Font(name=(MONO if c in (2, 3, 4, 5) else SANS),
                                          size=9.5, color=INK if c == 7 else MUTED)
        for c in range(1, 8):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = GRID
        ws.row_dimensions[r].height = max(40, 13 * (len(row[6]) // 58 + 1))
        r += 1
    ws.cell(r + 1, 1,
            "Bottom line: this is a pre-foliage trip. Mt Tsurugi's summit (1,955 m) is the "
            "one place turning while you're there — everything at sea level peaks in late "
            "November. Miyajima was recorded completely green on 12 October 2025.")\
      .font = Font(name=SANS, size=10, bold=True, color=ACCENT)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=7)
    ws.cell(r + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r + 1].height = 44
    ws.sheet_view.showGridLines = False

# ── Places tab: every stop, how long people spend, map link ─────────────────
def build_places(ws, opt):
    head = ["Date", "Day", "The day", "Place", "How long people spend",
            "Open in Google Maps", "Who says so / basis"]
    _title_block(ws, "Every place on this route — one row each, every pin live",
                 "THIS IS WHERE THE MAP PINS LIVE. A spreadsheet cell holds only one "
                 "hyperlink, so a day with nine places cannot carry nine pins on its "
                 "itinerary row \u2014 they get a row each here instead. Filter column A "
                 "by date to get one day. Durations are what named travellers or the "
                 "operator reported; where nobody publishes one, it says so.", 7)
    _header(ws, 4, head, wraps=(3, 7))
    _w(ws, [11, 6, 34, 38, 22, 26, 66])
    dwell = {d[0].lower(): d for d in DWELL}

    def match(place):
        p = place.lower()
        for k, d in dwell.items():
            key = k.split("(")[0].strip()
            if key in p or p in key or (len(p) > 6 and p[:9] in k):
                return d
        return None

    r, band = 5, 0
    for day in opt["days"]:
        dt = datetime.date.fromisoformat(day["date"])
        for name, url in day_places(day, opt["key"]):
            d = match(name)
            ws.cell(r, 1, day["date"]).font = Font(name=MONO, size=9.5, color=MUTED)
            ws.cell(r, 2, dt.strftime("%a")).font = Font(name=MONO, size=9.5, color=MUTED)
            ws.cell(r, 3, day["title"]).font = Font(name=SANS, size=9, color=MUTED)
            ws.cell(r, 4, name).font = Font(name=SANS, size=10.5, bold=True, color=INK)
            ws.cell(r, 5, d[1] if d else "not published")
            ws.cell(r, 5).font = Font(name=MONO, size=10, bold=bool(d),
                                      color=INK if d else MUTED)
            # the label carries the place name, so a filtered view still reads
            _link(ws, r, 6, f"\U0001f4cd {name} \u2197", url, size=9.5)
            if d:
                ws.cell(r, 7, d[3])
                ws.cell(r, 7).font = Font(name=SANS, size=9, color=MUTED)
            _band(ws, r, 7, band % 2 == 1)
            for c in range(1, 8):
                ws.cell(r, c).border = GRID
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = max(30, 12 * ((len(d[3]) // 74 + 1) if d else 1) + 14)
            r += 1
            band += 1
    ws.auto_filter.ref = f"A4:G{r-1}"
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

# ── Costs tab ───────────────────────────────────────────────────────────────
def build_costs(ws):
    head = ["What you pay for", "Price (Yen)", "Price (NIS)", "Source", "How solid", "Notes"]
    _title_block(ws, "Admissions, lifts and everything that isn't a transport fare",
                 "Transport fares live on the Transport tab. Every figure here is linked "
                 "to where it came from.", 6)
    ws.cell(4, 1, "Live rate  JPY → NIS").font = Font(name=SANS, size=9.5, bold=True)
    rate = ws.cell(4, 2, LIVE_FORMULA)
    rate.font = Font(name=MONO, size=11, bold=True, color=ACCENT)
    rate.number_format = "0.00000"
    rate.fill = PatternFill("solid", fgColor=ACC_BG)
    rate.alignment = Alignment(horizontal="center")
    ws.cell(3, 3, "Rate below is live in Google Sheets, fixed elsewhere — see the note")\
      .font = Font(name=SANS, size=8.5, italic=True, color=MUTED)
    ws.row_dimensions[3].height = 13
    ws.row_dimensions[5].height = 6
    _header(ws, 6, head, wraps=(6,))
    _w(ws, [40, 14, 15, 48, 13, 56])
    tier_colour = {"operator": "FF26624A", "official": "FF26624A",
                   "third-party": "FF8A6410", "traveller": "FF8A6410"}
    r = 7
    for i, (what, yen, key, note) in enumerate(ADMISSIONS):
        ws.cell(r, 1, what).font = Font(name=SANS, size=10.5, bold=True, color=INK)
        c = ws.cell(r, 2, yen); c.font = Font(name=MONO, size=10); c.number_format = '"¥"#,##0'
        c = ws.cell(r, 3, f"=ROUND(B{r}*$B$4,2)")
        c.font = Font(name=MONO, size=10, color=MUTED); c.number_format = '"₪"#,##0.00'
        ws.cell(r, 2).alignment = Alignment(horizontal="right", indent=1)
        c.alignment = Alignment(horizontal="right", indent=2)
        label, url, tier = SOURCES[key]
        _link(ws, r, 4, label, url, size=9)
        ws.cell(r, 5, tier).font = Font(name=SANS, size=9, bold=True, color=tier_colour[tier])
        ws.cell(r, 6, note).font = Font(name=SANS, size=9, color=MUTED)
        _band(ws, r, 6, i % 2 == 1)
        for cc in range(1, 7):
            ws.cell(r, cc).border = GRID
            ws.cell(r, cc).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 34
        r += 1
    ws.sheet_view.showGridLines = False

# ── Attractions tab ─────────────────────────────────────────────────────────
def build_attractions(ws):
    head = ["Attraction", "Where", "Adult price", "Opening hours",
            "Closed / seasonal", "Do you need to book?", "Source"]
    _title_block(ws, "What each place costs and when it's open",
                 "Adult prices, per person. Anything the operator doesn't publish says "
                 "\u201cnot stated\u201d rather than being guessed at.", 7)
    _header(ws, 4, head, wraps=(4, 5, 6))
    _w(ws, [38, 20, 34, 40, 34, 44, 40])
    r = 5
    for i, (name, where, price, hours, closed, book, key) in enumerate(ATTRACTIONS):
        u = pin(name) or pin(where)
        if u:
            _link(ws, r, 1, name, u, size=10.5, bold=True)
        else:
            ws.cell(r, 1, name).font = Font(name=SANS, size=10.5, bold=True, color=INK)
        ws.cell(r, 2, where).font = Font(name=SANS, size=9.5, color=MUTED)
        free = price.upper().startswith("FREE")
        ws.cell(r, 3, price).font = Font(name=MONO, size=9.5, bold=True,
                                         color=("FF26624A" if free else INK))
        ws.cell(r, 4, hours).font = Font(name=SANS, size=9.5, color=INK)
        ws.cell(r, 5, closed).font = Font(name=SANS, size=9.5, color=MUTED)
        must = book.upper().startswith("YES")
        ws.cell(r, 6, book).font = Font(name=SANS, size=9.5, bold=must,
                                        color=(ACCENT if must else MUTED))
        label, url, _t = SOURCES[key]
        _link(ws, r, 7, label, url, size=8.5)
        _band(ws, r, 7, i % 2 == 1)
        for c in range(1, 8):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = GRID
        ws.row_dimensions[r].height = max(32, 12 * (max(len(hours), len(book)) // 40 + 2))
        r += 1
    ws.auto_filter.ref = f"A4:G{r-1}"
    ws.sheet_view.showGridLines = False

# ── Bookings tab ────────────────────────────────────────────────────────────
def build_bookings(ws):
    _title_block(ws, "Book these in advance \u2014 everything else you can walk up to",
                 "The rest of the trip needs no reservations at all.", 4)
    _header(ws, 4, ["What", "How it works", "When to book", "Source"], wraps=(2, 3))
    _w(ws, [44, 50, 62, 44])
    r = 5
    for i, (what, how, when, key) in enumerate(BOOKINGS):
        ws.cell(r, 1, what).font = Font(name=SANS, size=11, bold=True, color=INK)
        ws.cell(r, 2, how).font = Font(name=SANS, size=10, color=INK)
        ws.cell(r, 3, when).font = Font(name=SANS, size=10, color=ACCENT)
        label, url, _t = SOURCES[key]
        _link(ws, r, 4, label, url, size=9)
        _band(ws, r, 4, i % 2 == 1)
        for c in range(1, 5):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = GRID
        ws.row_dimensions[r].height = max(40, 12 * (len(when) // 58 + 2))
        r += 1
    ws.sheet_view.showGridLines = False

# ── Tides tab ───────────────────────────────────────────────────────────────
def build_alternatives(ws, opt):
    """Days that can be spent more than one way.

    The dashboard swaps these in place on the day card; a spreadsheet cannot, so
    each alternative gets its own block here and the Itinerary row points at it.
    """
    head = ["Date", "Instead of", "This alternative", "What you'll actually do",
            "Getting there", "Watch out for", "Legs", "Cost, one person"]
    days = [d for d in opt["days"] if d.get("alts")]
    _title_block(ws, "Days you can spend more than one way",
                 "Each block replaces the WHOLE day — highlights, travel and costs together. "
                 "Pick one; don't mix. Costs are per person and cover transport only.", 8)
    _header(ws, 4, head, wraps=(4, 5, 6))
    _w(ws, [11, 26, 34, 68, 42, 52, 8, 16])
    r = 5
    if not days:
        ws.cell(r, 1, "No day in this option has alternatives.")\
          .font = Font(name=SANS, size=10, italic=True, color=MUTED)
        return
    for d in days:
        blocks = [dict(label=d.get("alt_label", "As planned"), do=d["do"],
                       travel=d["travel"], watch=d["watch"], legs=d["legs"])] + list(d["alts"])
        for j, b in enumerate(blocks):
            cost = sum(l[5] for l in b["legs"] if isinstance(l[5], (int, float)))
            ws.cell(r, 1, d["date"]).font = Font(name=MONO, size=9.5, color=MUTED)
            ws.cell(r, 2, d["title"]).font = Font(name=SANS, size=9.5, color=MUTED)
            c3 = ws.cell(r, 3, ("DEFAULT \u2014 " if j == 0 else "") + b["label"])
            c3.font = Font(name=SANS, size=10.5, bold=True,
                           color=(INK if j == 0 else ACCENT))
            ws.cell(r, 4, _highlights(b))
            ws.cell(r, 5, b["travel"])
            ws.cell(r, 6, _bullets(b["watch"]))
            ws.cell(r, 7, len(b["legs"])).font = Font(name=MONO, size=9.5)
            cc = ws.cell(r, 8, f"\u00a5{cost:,}" if cost else "\u2014")
            cc.font = Font(name=MONO, size=10, bold=True, color=INK)
            _band(ws, r, 8, j % 2 == 1)
            for c in range(1, 9):
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(r, c).border = GRID
            ws.row_dimensions[r].height = max(74, 11.5 * (len(_highlights(b)) // 70 + 2))
            r += 1
        r += 1
    ws.freeze_panes = "A5"


def build_walks(ws, opt):
    """Every walk with its trailhead, the two GPS apps, and a blunt verdict.

    This table already existed in refs/trails.py and was rendered NOWHERE — the
    traveller asked twice for AllTrails/YAMAP cross-references and could not find
    them, because collecting the data and shipping it are two different jobs.
    """
    head = ["The walk", "Where", "Which day", "Trailhead \u2014 tap for the exact start",
            "Distance & ascent", "Time", "How hard", "AllTrails", "YAMAP",
            "What the description page doesn't tell you", "Source"]
    _title_block(ws, "Every walk, with the exact spot it starts from",
                 "Where a walk STARTS is the hardest thing to work out on the ground, so "
                 "each one is a tappable pin. AllTrails and YAMAP let your phone track you "
                 "offline; YAMAP has far better coverage of Japanese trails. Distance, "
                 "ascent and time are the app's figures or the operator's \u2014 never mine. "
                 "\u201cWhich day\u201d tells you where each one falls in THIS itinerary.", 11)
    _header(ws, 4, head, wraps=(3, 4, 10))
    _w(ws, [46, 20, 22, 26, 20, 20, 16, 22, 22, 72, 32])
    when = {}
    for day in opt["days"]:
        for i in walks_for(day):
            when.setdefault(i, []).append(day["date"][5:])
    r = 5
    for i, (name, where, ll, yamap, at, dist, time, verdict, note, key) in enumerate(TRAILHEADS):
        ws.cell(r, 1, name).font = Font(name=SANS, size=10.5, bold=True, color=INK)
        ws.cell(r, 2, where).font = Font(name=SANS, size=9.5, color=MUTED)
        days = when.get(i)
        ws.cell(r, 3, ", ".join(days) if days else "not on this itinerary")
        ws.cell(r, 3).font = Font(name=MONO, size=9.5, bold=bool(days),
                                  color=(ACCENT if days else MUTED))
        _link(ws, r, 4, ll, "https://www.google.com/maps/search/?api=1&query="
                            + ll.replace(",", "%2C"), size=9.5, bold=True)
        ws.cell(r, 5, dist).font = Font(name=MONO, size=9.5, color=INK)
        ws.cell(r, 6, time).font = Font(name=MONO, size=9.5, color=INK)
        hard = not verdict.upper().startswith("EASY")
        ws.cell(r, 7, verdict).font = Font(name=SANS, size=9.5, bold=True,
                                           color=(ACCENT if hard else "FF26624A"))
        if at:
            _link(ws, r, 8, "AllTrails route", at, size=9.5)
        else:
            ws.cell(r, 8, "no matching route").font = Font(name=SANS, size=9, italic=True, color=MUTED)
        if yamap:
            _link(ws, r, 9, "YAMAP route", yamap, size=9.5)
        else:
            ws.cell(r, 9, "none").font = Font(name=SANS, size=9, italic=True, color=MUTED)
        ws.cell(r, 10, note).font = Font(name=SANS, size=9.5, color=INK)
        label, url, _t = SOURCES[key]
        _link(ws, r, 11, label, url, size=8.5)
        _band(ws, r, 11, i % 2 == 1)
        for c in range(1, 12):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = GRID
        ws.row_dimensions[r].height = max(58, 11.5 * (len(note) // 80 + 2))
        r += 1
    ws.freeze_panes = "A5"


def build_tides(ws):
    head = ["What depends on the tide", "When, on this trip", "The rule", "Tide table"]
    _title_block(ws, f"{len(TIDES)} things here only work at the right water level",
                 "Check these BEFORE fixing the hour of the day. All are free to look up.", 4)
    _header(ws, 4, head, wraps=(3,))
    _w(ws, [34, 34, 82, 52])
    r = 5
    for i, (what, when, rule, key, extra) in enumerate(TIDES):
        ws.cell(r, 1, what).font = Font(name=SANS, size=11, bold=True, color=INK)
        ws.cell(r, 2, when).font = Font(name=MONO, size=9.5, color=ACCENT)
        ws.cell(r, 3, rule).font = Font(name=SANS, size=10, color=INK)
        label, url, _t = SOURCES[key]
        _link(ws, r, 4, label, url, size=9.5, bold=True)
        _band(ws, r, 4, i % 2 == 1)
        for c in range(1, 5):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, c).border = GRID
        ws.row_dimensions[r].height = max(56, 12 * (len(rule) // 76 + 2))
        r += 1
        for ek in extra:
            el, eu, _ = SOURCES[ek]
            ws.cell(r, 3, "also:").font = Font(name=SANS, size=9, italic=True, color=MUTED)
            _link(ws, r, 4, el, eu, size=9)
            _band(ws, r, 4, i % 2 == 1)
            for c in range(1, 5):
                ws.cell(r, c).border = GRID
                ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 20
            r += 1
    ws.sheet_view.showGridLines = False

# ── Sources tab ─────────────────────────────────────────────────────────────
def build_sources(ws):
    head = ["Source", "How solid", "Link"]
    _title_block(ws, "Every source behind this plan",
                 "operator = the company that runs it  ·  official = a government or "
                 "tourism body  ·  third-party = a fare database or guidebook  ·  "
                 "traveller = someone who actually paid it", 3)
    _header(ws, 4, head)
    _w(ws, [72, 15, 88])
    tier_colour = {"operator": "FF26624A", "official": "FF26624A",
                   "third-party": "FF8A6410", "traveller": "FF8A6410"}
    order = {"operator": 0, "official": 1, "third-party": 2, "traveller": 3}
    rows = sorted(SOURCES.values(), key=lambda v: (order[v[2]], v[0].lower()))
    r = 5
    for i, (label, url, tier) in enumerate(rows):
        ws.cell(r, 1, label).font = Font(name=SANS, size=10, color=INK)
        ws.cell(r, 2, tier).font = Font(name=SANS, size=9.5, bold=True, color=tier_colour[tier])
        _link(ws, r, 3, url, url, size=9)
        _band(ws, r, 3, i % 2 == 1)
        for c in range(1, 4):
            ws.cell(r, c).border = GRID
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[r].height = 22
        r += 1
    ws.sheet_view.showGridLines = False

# ── drive ───────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUT, exist_ok=True)
    for key, opt in OPTIONS.items():
        wb = Workbook()
        wb.remove(wb.active)
        for name, fn, colour in (
            ("Itinerary", lambda ws: build_itinerary(ws, opt), TAB_A),
            ("Transport", lambda ws: build_transport(ws, opt), TAB_B),
            ("Places & timings", lambda ws: build_places(ws, opt), "3E6B8A"),
            ("Walks & trailheads", lambda ws: build_walks(ws, opt), "3E7A4E"),
            ("Costs",     build_costs,    "8A5B3E"),
            ("Attractions", build_attractions, "7A4E8A"),
            ("Book ahead", build_bookings,   "A33B2E"),
            ("Glossary",  build_glossary, "8A9099"),
            ("Foliage",   build_foliage,  "B08A3E"),
            ("Day alternatives", lambda ws: build_alternatives(ws, opt), "6E4E8A"),
            ("Tides",     build_tides,    "2F6E8F"),
            ("Sources",   build_sources,  "5A5A5A"),
        ):
            ws = wb.create_sheet(name)
            ws.sheet_properties.tabColor = colour
            fn(ws)
        wb.active = 0
        path = os.path.join(OUT, f"Shikoku-Option-{key}.xlsx")
        wb.save(path)

        # CSVs, quoted properly this time
        with open(os.path.join(OUT, f"option{key.lower()}-itinerary.csv"),
                  "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, quoting=csv.QUOTE_ALL)
            w.writerow(ITIN_HEAD)
            w.writerows(itinerary_rows(opt))
        with open(os.path.join(OUT, f"option{key.lower()}-transport.csv"),
                  "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f, quoting=csv.QUOTE_ALL)
            w.writerow(TR_HEAD)
            for row in transport_rows(opt):
                src = resolve(row[5])
                if row[6] is None:
                    ref, tier = "no published figure found", "—"
                elif src:
                    ref, tier = f"{src[0]} — {src[1]}", src[2]
                else:
                    ref, tier = "unsourced — verify", "check"
                w.writerow([("" if v is None else v) for v in row] + ["", ref, tier])

        days = len(opt["days"])
        legs = len(transport_rows(opt))
        total = sum(l[5] for d in opt["days"] for l in d["legs"] if l[5] is not None)
        print(f"Option {key}: {days} days, {legs} legs, ¥{total:,}  → {os.path.basename(path)}")

if __name__ == "__main__":
    main()
